#!/usr/bin/env python3

import sys
import json
import os
import shutil
from box import Box
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Fill
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle

# Estilos para celdas
numberStyle = NamedStyle(name='numberStyle', number_format='0.00')

# Importamos Jinja2
import jinja2

# Cargar el JSON desde un archivo

if sys.argv[1] == "DAW":

    with open('./boe/rd-daw.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

data_box = Box(data)

# Abrimos el excel
#writer = pd.ExcelWriter("datos.xlsx", engine="openpyxl")

libro="PDFS/"+str(sys.argv[1])+"_libro.xlsx"
wb = openpyxl.Workbook()

# Algunas posiciones fijas
p_codigo='C1'
p_nombre='C2'
p_ra_col_l='B'
p_ra_titulo_col=2
p_ra_titulo_row=8
p_ce_col_l='E'
p_req_fe_col_l='H'
p_contenidos_col_l='J'

p_TOTAL_HORAS_titulo="F3"
p_TOTAL_HORAS="F5"
p_TOTAL_HORAS_DUAL_titulo="I3"
p_TOTAL_HORAS_DUAL="I5"


for codigo in data_box.ModulosProfesionales:

    modulo=data_box.ModulosProfesionales[codigo]
    print(modulo.nombre)

    p_ra_titulo_col=2
    p_ra_titulo_row=8

    ws = wb.create_sheet(title=modulo.nombre)
    wb.active = wb.sheetnames.index(modulo.nombre)

    ws['B1'].value="Código"
    ws['B1'].alignment = Alignment(horizontal='center',vertical='center')
    ws['B1'].fill = PatternFill('lightHorizontal')

    ws['B2'].value="Nombre"
    ws['B2'].alignment = Alignment(horizontal='center',vertical='center')
    ws['B2'].fill = PatternFill('lightHorizontal')

    ws.merge_cells(start_row=1,start_column=3,end_row=1,end_column=5)
    ws[p_codigo].value=codigo
    ws[p_codigo].alignment = Alignment(horizontal='center',vertical='center')
    ws[p_codigo].fill = PatternFill('darkTrellis')
    ws[p_codigo].font = Font(size=13)

    ws.merge_cells(start_row=2,start_column=3,end_row=2,end_column=5)
    ws[p_nombre].value=modulo.nombre
    ws[p_nombre].alignment = Alignment(horizontal='center',vertical='center')
    ws[p_nombre].fill = PatternFill('darkTrellis')
    ws[p_nombre].font = Font(size=14)

    ws.merge_cells(start_row=3,start_column=6,end_row=4,end_column=6)
    ws[p_TOTAL_HORAS_titulo].value="TOTAL HORAS"
    ws[p_TOTAL_HORAS_titulo].fill = PatternFill('darkTrellis')
    ws[p_TOTAL_HORAS_titulo].font = Font(size=13)
    ws[p_TOTAL_HORAS_titulo].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws.column_dimensions['I'].width = 15

    ws[p_TOTAL_HORAS].font = Font(size=16)

    # TODO
    ws[p_TOTAL_HORAS].value="=SUM(F8:F200)/2"

    ws.merge_cells(start_row=3,start_column=9,end_row=4,end_column=9)
    ws[p_TOTAL_HORAS_DUAL_titulo].value="TOTAL H.DUAL"
    ws[p_TOTAL_HORAS_DUAL_titulo].fill = PatternFill('darkTrellis')
    ws[p_TOTAL_HORAS_DUAL_titulo].font = Font(size=13)
    ws[p_TOTAL_HORAS_DUAL_titulo].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws[p_TOTAL_HORAS_DUAL].font = Font(size=14)
    ws.column_dimensions['F'].width = 15
    # TODO
    ws[p_TOTAL_HORAS_DUAL].value="=SUM(I8:I200)/2"



    print(" - Resultados de Aprendizaje ")
    ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_ra_titulo_col, end_row=p_ra_titulo_row+1, end_column=p_ra_titulo_col)
    ws.cell(column=p_ra_titulo_col,row=p_ra_titulo_row).value="RESULTADO DE APRENDIZAJE"
    ws.cell(column=p_ra_titulo_col,row=p_ra_titulo_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_ra_titulo_col,row=p_ra_titulo_row).fill = PatternFill('gray125')

    ws.column_dimensions[p_ra_col_l].width = 40

    print(" - %RA ")
    p_percent_ra_col=p_ra_titulo_col+1
    p_percent_ra_row=p_ra_titulo_row
    ws.merge_cells(start_row=p_percent_ra_row, start_column=p_percent_ra_col, end_row=p_percent_ra_row+1, end_column=p_percent_ra_col)
    ws.cell(column=p_percent_ra_col,row=p_percent_ra_row).value="% RA"
    ws.cell(column=p_percent_ra_col,row=p_percent_ra_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_percent_ra_col,row=p_percent_ra_row).fill = PatternFill('gray125')

    print(" - COMP ")
    p_comp_col=p_percent_ra_col+1
    p_comp_row=p_percent_ra_row
    ws.merge_cells(start_row=p_comp_row, start_column=p_comp_col, end_row=p_comp_row+1, end_column=p_comp_col)
    ws.cell(column=p_comp_col,row=p_comp_row).value="COMP"
    ws.cell(column=p_comp_col,row=p_comp_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_comp_col,row=p_comp_row).fill = PatternFill('gray125')

    print(" - CRITERIOS DE EVALUACIÓN ")
    p_ce_col=p_comp_col+1
    p_ce_row=p_comp_row
    ws.merge_cells(start_row=p_ce_row, start_column=p_ce_col, end_row=p_comp_row+1, end_column=p_ce_col)
    ws.cell(column=p_ce_col,row=p_ce_row).value="CRITERIOS DE EVALUACIÓN"
    ws.cell(column=p_ce_col,row=p_ce_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_ce_col,row=p_ce_row).fill = PatternFill('gray125')

    ws.column_dimensions[p_ce_col_l].width = 90



    print(" - HORAS ")
    p_h_col=p_ce_col+1
    p_h_row=p_ce_row
    ws.merge_cells(start_row=p_h_row, start_column=p_h_col, end_row=p_h_row+1, end_column=p_h_col)
    ws.cell(column=p_h_col,row=p_h_row).value="HORAS"
    ws.cell(column=p_h_col,row=p_h_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_h_col,row=p_h_row).fill = PatternFill('gray125')


    print(" - %CE ")
    p_ce_per_col=p_h_col+1
    p_ce_per_row=p_h_row
    ws.merge_cells(start_row=p_ce_per_row, start_column=p_ce_per_col, end_row=p_ce_per_row+1, end_column=p_ce_per_col)
    ws.cell(column=p_ce_per_col,row=p_ce_per_row).value="% CE"
    ws.cell(column=p_ce_per_col,row=p_ce_per_row).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(column=p_ce_per_col,row=p_ce_per_row).fill = PatternFill('gray125')

    print(" - REQUISITO FE")
    p_req_fe_col=p_ce_per_col+1
    p_req_fe_row=p_h_row
    ws.merge_cells(start_row=p_req_fe_row, start_column=p_req_fe_col, end_row=p_req_fe_row+1, end_column=p_req_fe_col)
    ws.cell(column=p_req_fe_col,row=p_req_fe_row).value="REQUISITO FE"
    ws.cell(column=p_req_fe_col,row=p_req_fe_row).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws.cell(column=p_req_fe_col,row=p_req_fe_row).fill = PatternFill('gray125')
    ws.column_dimensions[p_req_fe_col_l].width =15

    print(" - HORAS DUAL ")
    p_horas_dual_col=p_req_fe_col+1
    p_horas_dual_row=p_req_fe_row
    ws.merge_cells(start_row=p_horas_dual_row, start_column=p_horas_dual_col, end_row=p_horas_dual_row+1, end_column=p_horas_dual_col)
    ws.cell(column=p_horas_dual_col,row=p_horas_dual_row).value="HORAS DUAL"
    ws.cell(column=p_horas_dual_col,row=p_horas_dual_row).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws.cell(column=p_horas_dual_col,row=p_horas_dual_row).fill = PatternFill('gray125')


    print(" - CONTENIDOS ")
    p_contenidos_col=p_horas_dual_col+1
    p_contenidos_row=p_horas_dual_row
    ws.merge_cells(start_row=p_contenidos_row, start_column=p_contenidos_col, end_row=p_contenidos_row+1, end_column=p_contenidos_col)
    ws.cell(column=p_contenidos_col,row=p_contenidos_row).value="CONTENIDOS"
    ws.cell(column=p_contenidos_col,row=p_contenidos_row).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
    ws.cell(column=p_contenidos_col,row=p_contenidos_row).fill = PatternFill('gray125')

    ws.column_dimensions[p_contenidos_col_l].width = 50


    # Ahora ya seguimos para abajo
    p_ra_titulo_row=p_ra_titulo_row+2
    ra_per=100/len(modulo.ResultadosAprendizaje)


    for ra in modulo.ResultadosAprendizaje:
        listaCriterios=modulo.ResultadosAprendizaje[ra].CriteriosEvaluacion
        numCriterios=len(listaCriterios)

        ws.cell(column=p_ra_titulo_col,row=p_ra_titulo_row).value=ra+"."+modulo.ResultadosAprendizaje[ra].Resultado
        ws.cell(column=p_ra_titulo_col,row=p_ra_titulo_row).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_ra_titulo_col, end_row=p_ra_titulo_row+numCriterios+1, end_column=p_ra_titulo_col)

        ws.cell(column=p_ra_titulo_col+1,row=p_ra_titulo_row).value=ra_per
        ws.cell(column=p_ra_titulo_col+1,row=p_ra_titulo_row).style = numberStyle
        ws.cell(column=p_ra_titulo_col+1,row=p_ra_titulo_row).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_ra_titulo_col+1, end_row=p_ra_titulo_row+numCriterios+1, end_column=p_ra_titulo_col+1)

        # Contenidos
        ws.merge_cells(start_row=p_ra_titulo_row, start_column=p_contenidos_col, end_row=p_ra_titulo_row+numCriterios, end_column=p_contenidos_col)
        # Criterios de Evaluacion de cada uno de los RA

        # TODOS
        p_ce_row=p_ce_row+2
        ws.cell(column=p_ce_col,row=p_ce_row).value="TODOS"
        ws.cell(column=p_ce_col,row=p_ce_row).alignment=Alignment(horizontal='center', vertical='center')

        ws.cell(column=p_ce_col+1,row=p_ce_row).value="=SUM(F"+str(p_ce_row+1)+":F"+str(p_ce_row+numCriterios)+")"
        ws.cell(column=p_ce_col+2,row=p_ce_row).value="=SUM(G"+str(p_ce_row+1)+":G"+str(p_ce_row+numCriterios)+")"
        ws.cell(column=p_ce_col+4,row=p_ce_row).value="=SUM(I"+str(p_ce_row+1)+":I"+str(p_ce_row+numCriterios)+")"

        ws.cell(column=p_ce_col,row=p_ce_row).fill = PatternFill('gray0625')
        ws.cell(column=p_ce_col+1,row=p_ce_row).fill = PatternFill('gray0625')
        ws.cell(column=p_ce_col+2,row=p_ce_row).fill = PatternFill('gray0625')
        ws.cell(column=p_ce_col+4,row=p_ce_row).fill = PatternFill('gray0625')

        # Ingenieria para las competencias
        if numCriterios < 3:
            print(" * Demasiados pocos criterios, no se hace nada mas que los titulos")
            ws.cell(column=p_comp_col,row=p_ce_row+1).value="CPROF"
            ws.cell(column=p_comp_col,row=p_ce_row+2).value="EMPLEA"
        else:
            if numCriterios % 2 == 1:
                numCProf = numCriterios // 2 + 1
            else:
                numCProf = numCriterios // 2
            numCEmplea = numCriterios - numCProf
            print(" -- NCRITERIOS "+str(numCriterios))
            print(" -- Tenemos CPROF : "+str(numCProf))
            print(" -- Tenemos EMPLEA : "+str(numCEmplea))

            ws.cell(column=p_comp_col,row=p_ce_row).value="CPROF"
            ws.cell(column=p_comp_col,row=p_ce_row).fill = PatternFill('gray125')
            ws.merge_cells(start_row=p_ce_row+1, start_column=p_comp_col, end_row=p_ce_row+numCProf, end_column=p_comp_col)
            ws.cell(column=p_comp_col,row=p_ce_row+1).alignment=Alignment(horizontal='center', vertical='center')

            ws.cell(column=p_comp_col,row=p_ce_row+numCProf+1).value="EMPLEA"
            ws.cell(column=p_comp_col,row=p_ce_row+numCProf+1).fill = PatternFill('gray125')
            ws.merge_cells(start_row=p_ce_row+2+numCProf, start_column=p_comp_col, end_row=p_ce_row+2+numCProf+numCEmplea-1, end_column=p_comp_col)
            ws.cell(column=p_comp_col,row=p_ce_row+numCProf+2).alignment=Alignment(horizontal='center', vertical='center')


        ce_per=100/numCriterios

        for ce in listaCriterios:
            p_ce_row=p_ce_row+1
            ws.cell(column=p_ce_col,row=p_ce_row).value=ce+") "+listaCriterios[ce]
            ws.cell(column=p_ce_col,row=p_ce_row).alignment = Alignment(horizontal='left', vertical='center',wrap_text=True)
            # Horas del Criterio
            ws.cell(column=p_ce_col+1,row=p_ce_row).value=0
            # Porcentaje del Criterio
            ws.cell(column=p_ce_col+2,row=p_ce_row).value=ce_per
            ws.cell(column=p_ce_col+2,row=p_ce_row).style = numberStyle
            ws.cell(column=p_ce_col+2,row=p_ce_row).alignment = Alignment(horizontal='right', vertical='center',wrap_text=True)

        # Incrementamos la fila
        p_ra_titulo_row=p_ra_titulo_row+numCriterios+2


print(" * Quitamos la primera hoja ")
del wb['Sheet']
print(" * Guardamos el libro en : "+str(libro))
wb.save(libro)

